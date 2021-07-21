import re
from tqdm import tqdm
from enum import Enum
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill

class Mode(Enum):
  SAFE=1
  ALL=2

def get_cell_value(cell):
  sheet = cell.parent
  if not isinstance(cell, MergedCell):
    return cell.value, cell.coordinate
  
  for range in sheet.merged_cells.ranges:
    if cell.coordinate in range:
      return range.start_cell.value, range.start_cell.coordinate
  raise AssertionError('Merged cell is not in any merge range!')

def is_price(value):
  return value is not None and len(value) > 2 and re.match(r'^[+-]?[0-9]{1,3}(?:\.[0-9]{3})*(?:,[0-9]{1,2})?$', value)

def load_worksheet(input):
  wb = load_workbook(input)
  first_sheetname = wb.sheetnames[0]
  ws = wb[first_sheetname]

  return wb, ws

def search_price_cell(cell, mode=Mode.SAFE):
  value = cell.value

  if mode == Mode.ALL:
    if type(value) == str and is_price(value):
      return [cell.coordinate]

    return []

  if type(value) != str or value not in ['VND', 'USD']:
    return []

  cell_list = []

  for i in range(-5, 5):
    try:
      neighbor_cell = cell.offset(row=0, column=i)
      neighbor_value, coord = get_cell_value(neighbor_cell)
      if is_price(neighbor_value):
        cell_list.append(coord)
        break
    except:
      pass
  
  return cell_list

def scan_price(ws, mode: Mode):
  coordinates = []
  total_size = (ws.max_row - ws.min_row) * (ws.max_column - ws.min_column)
  skip_keywords = ['Số lượng', 'trọng lượng hàng (Gross)']

  with tqdm(total=total_size) as pbar:
    for row in ws.rows:
      for cell in row:
        should_skip_this_row = False

        for keyword in skip_keywords:
          if type(cell.value) == str and keyword in cell.value:
            should_skip_this_row = True
            break

        if should_skip_this_row: 
          break
        coordinates.extend(search_price_cell(cell, mode))
        pbar.update(1)

  return coordinates

def remove_price(wb, ws, coordinates):
  for coord in coordinates:
    ws[coord].value = ''
  wb.save('output.xlsx')
  for coord in coordinates:
    ws[coord].fill = PatternFill("solid", fgColor="DDDDDD")
  wb.save('outputWithHighlight.xlsx')

def main(input):
  wb, ws = load_worksheet(input)
  coordinates = scan_price(ws, Mode.ALL)
  remove_price(wb, ws, coordinates)

if __name__ == '__main__':
  input = 'input.xlsx'
  main(input)